#!/usr/bin/env python
# -*- coding: utf-8 -*-
# @Author   : xiayun
# @Time     : 2022/4/7 19:50
# @Description     : 这是一个**功能的Python文件
# @Software : PyCharm
import os
import threading
import time

import openpyxl
from openpyxl.styles import Font, Color

from Keys_Excel.Log.logs import logger

log = logger()



def read_excel(file_path=None, _sheet_index=None, _have_title=True):
    """
    读取测试用例并执行
    :param file_path: 测试用例路径
    :param _sheet_index: sheet页索引
    :param _have_title: 是否含有标题
    :return: 返回每一行的测试数据
    """
    # 打开用例
    workbook = openpyxl.load_workbook(file_path)
    # 测试用例数据
    excel_data = {}
    # 默认获取所有的sheet页
    if _sheet_index is None:
        sheet_names = workbook.sheetnames
    else:
        # 否则获取对应的sheet页
        sheet_names = workbook.sheetnames[_sheet_index]
    # 循环获取所有的sheet页
    for i in sheet_names:
        sheet = workbook[i]
        # 获取所有数据
        if _have_title:
            data = list(sheet.values)[1:]
        else:
            data = sheet.values
        all_data = []
        for value in data:
            excel_data = {'param': {}}
            # 过滤掉第二行的的表头
            if isinstance(value[0], int):
                excel_data['id'] = value[0]
                excel_data['operation'] = value[1]
                # 增加判断，如果是断言，则把预期结果也加到参数中
                if 'assert' in value[1]:
                    excel_data['param']['txt'] = value[4]
                # 参数如果不为空
                if value[2]:
                    for param in value[2].split(';'):
                        key_ = param.split('=')[0]
                        value_ = param.split('=')[1]
                        excel_data['param'][key_] = value_
                excel_data['description'] = value[3]
                # 打印日志
                log.debug(excel_data['description'])
                all_data.append(excel_data)
        assert_Res = yield all_data  # 返回断言结果
        # 成功则回填Pass，失败则Failed
        for assert_data in all_data:
            if 'assert' in assert_data['operation']:
                if assert_Res == 'Pass':
                    sheet.cell(assert_data['id'] + 2, 6).value = "Pass"
                    sheet.cell(assert_data['id'] + 2, 6).font = Font(name='仿宋', size=12, color=Color(index=3), b=True)
                elif assert_Res == 'Failed':
                    sheet.cell(assert_data['id'] + 2, 6).value = "Failed"
                    sheet.cell(assert_data['id'] + 2, 6).font = Font(name='仿宋', size=12, color=Color(index=2), b=True)
        workbook.save(file_path)
